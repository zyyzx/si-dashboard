"""Data loaders for the SI candidate-generation pipeline.

Reads source files read-only and returns tidy pandas DataFrames /
typed dicts. No mutations to upstream data, no network calls.

Public surface:
    load_si_history()        -> DataFrame: ticker, date, si, dtc, ...
    load_equities()          -> DataFrame: ticker, name, sector, market_cap_bucket, ...
    load_float_panel()       -> DataFrame: ticker, asof_date, float_shares
    load_joined_si_with_meta()-> DataFrame: SI history + sector + float (point-in-time)
"""

from __future__ import annotations

import datetime as _dt
from pathlib import Path
from typing import Optional

import pandas as pd

from . import (
    SI_HISTORY_CSV,
    EQUITIES_CSV,
    CAPIQ_FLOAT_XLSX,
    FLOAT_PANEL_PARQUET,
)


# ----------------------------- SI history -----------------------------

# FINRA columns we keep. Drop noise columns aggressively to keep
# the feature table small.
_SI_KEEP_COLS = {
    "symbolCode": "ticker",
    "issueName": "name",
    "issuerServicesGroupExchangeCode": "exchange",
    "marketClassCode": "market_class",
    "currentShortPositionQuantity": "si_shares",
    "previousShortPositionQuantity": "si_shares_prev",
    "averageDailyVolumeQuantity": "adv_shares",
    "daysToCoverQuantity": "dtc",
    "settlementDate": "settlement_date",
}


def load_si_history(path: Path = SI_HISTORY_CSV) -> pd.DataFrame:
    """Load the canonical FINRA short-interest history.

    Returns one row per (ticker, settlement_date) with numeric coercion
    and a parsed datetime ``settlement_date`` column. ``dtc`` of FINRA's
    sentinel value 999.99 is preserved (it means "no avg-volume data");
    callers should filter as appropriate.
    """
    df = pd.read_csv(
        path,
        usecols=list(_SI_KEEP_COLS.keys()),
        dtype={"settlementDate": str, "symbolCode": str, "issueName": str},
        low_memory=False,
    )
    df = df.rename(columns=_SI_KEEP_COLS)

    # numeric coercion (errors → NaN)
    for c in ("si_shares", "si_shares_prev", "adv_shares", "dtc"):
        df[c] = pd.to_numeric(df[c], errors="coerce")

    df["settlement_date"] = pd.to_datetime(df["settlement_date"], format="%Y%m%d", errors="coerce")
    df = df.dropna(subset=["ticker", "settlement_date", "si_shares"])
    df["ticker"] = df["ticker"].str.upper().str.strip()

    # de-dupe in case a (ticker, date) appears twice (FINRA revisions)
    df = (
        df.sort_values(["ticker", "settlement_date"])
          .drop_duplicates(subset=["ticker", "settlement_date"], keep="last")
          .reset_index(drop=True)
    )
    return df


# ----------------------------- Equities meta --------------------------

_EQUITIES_KEEP = {
    "symbol": "ticker",
    "name": "name_eq",
    "sector": "sector",
    "industry_group": "industry_group",
    "industry": "industry",
    "exchange": "exchange_eq",
    "market": "market",
    "country": "country",
    "market_cap": "market_cap_bucket",
}


def load_equities(path: Path = EQUITIES_CSV) -> pd.DataFrame:
    """Load equity universe metadata.

    Filters to US listings only (NYSE / NASDAQ / NYSE American / OTC) so
    sector medians are computed against a sensible peer set, not against
    Shenzhen-listed companies.

    ``market_cap`` in equities.csv is a bucket label (e.g. "Large Cap",
    "Mid Cap", "Small Cap", "Micro Cap"), not a numeric value — we
    expose it as ``market_cap_bucket``. A numeric market cap will need
    to be sourced from prices × shares in Phase 1B.
    """
    df = pd.read_csv(
        path,
        usecols=list(_EQUITIES_KEEP.keys()),
        dtype=str,
        low_memory=False,
    )
    df = df.rename(columns=_EQUITIES_KEEP)
    df["ticker"] = df["ticker"].str.upper().str.strip()

    us_markets = {
        "New York Stock Exchange",
        "NASDAQ Stock Exchange",
        "NYSE American",
        "OTC Markets",
        "BATS Global Markets",
        "Cboe BZX U.S. Equities Exchange",
    }
    df = df[df["market"].isin(us_markets) | df["country"].eq("United States")]

    # one row per ticker (in case of dupes); keep the row with most non-null cells
    df = (
        df.assign(_filled=lambda x: x.notna().sum(axis=1))
          .sort_values("_filled", ascending=False)
          .drop_duplicates(subset=["ticker"], keep="first")
          .drop(columns="_filled")
          .reset_index(drop=True)
    )
    return df


# ----------------------------- Float panel ----------------------------

def _load_capiq_historical_raw(path: Path = CAPIQ_FLOAT_XLSX) -> pd.DataFrame:
    """Read the CapIQ historical float Excel into long-form.

    Mirrors the loader in integrate_float_data.py but returns a tidy
    DataFrame instead of a nested dict. CapIQ values come in millions;
    we convert to raw shares (× 1e6).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Float"]

    # Header row → list of date strings (from datetime cells or stringified)
    dates: list[str] = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        if isinstance(val, _dt.datetime):
            dates.append(val.strftime("%Y-%m-%d"))
        elif isinstance(val, str):
            s = val.replace("/", "-")
            # accept either YYYY-MM-DD or YYYYMMDD
            if len(s) == 8 and s.isdigit():
                s = f"{s[:4]}-{s[4:6]}-{s[6:]}"
            dates.append(s)
        else:
            dates.append(str(val))

    rows: list[tuple[str, str, float]] = []
    for r in range(2, ws.max_row + 1):
        ticker = ws.cell(row=r, column=1).value
        if not ticker:
            continue
        ticker = str(ticker).strip().upper()
        for j, d in enumerate(dates):
            cell = ws.cell(row=r, column=j + 2).value
            if isinstance(cell, (int, float)) and cell > 0:
                rows.append((ticker, d, float(cell) * 1e6))
    wb.close()

    out = pd.DataFrame(rows, columns=["ticker", "asof_date", "float_shares"])
    out["asof_date"] = pd.to_datetime(out["asof_date"], errors="coerce")
    out = out.dropna(subset=["asof_date"]).reset_index(drop=True)
    return out


def load_float_panel(
    rebuild: bool = False,
    cache_path: Path = FLOAT_PANEL_PARQUET,
) -> pd.DataFrame:
    """Load the float panel, caching the parsed CapIQ Excel as parquet.

    The Excel parse is slow (~30s for 11MB / 13.8K tickers × 26 dates),
    so we cache it. Pass ``rebuild=True`` to force a fresh parse.

    Returns columns: ticker, asof_date (datetime), float_shares (raw).
    """
    if cache_path.exists() and not rebuild:
        return pd.read_parquet(cache_path)

    if not CAPIQ_FLOAT_XLSX.exists():
        # No float source available — return empty panel with right schema
        return pd.DataFrame({
            "ticker": pd.Series(dtype=str),
            "asof_date": pd.Series(dtype="datetime64[ns]"),
            "float_shares": pd.Series(dtype=float),
        })

    df = _load_capiq_historical_raw(CAPIQ_FLOAT_XLSX)
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_parquet(cache_path, index=False)
    return df


# ----------------------------- Joined view ----------------------------

def load_joined_si_with_meta(
    si: Optional[pd.DataFrame] = None,
    eq: Optional[pd.DataFrame] = None,
    floats: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """Return SI history left-joined with sector metadata and
    point-in-time float (forward-filled per ticker).

    Output columns: ticker, settlement_date, si_shares, si_shares_prev,
    adv_shares, dtc, name, exchange, market_class, sector,
    industry_group, industry, market_cap_bucket, float_shares,
    si_pct_float.

    Only US-listed tickers are retained (via the equities filter).
    """
    si = load_si_history() if si is None else si
    eq = load_equities() if eq is None else eq
    floats = load_float_panel() if floats is None else floats

    # Inner-join on ticker keeps only US-listed names with FINRA SI
    df = si.merge(
        eq[["ticker", "sector", "industry_group", "industry", "market_cap_bucket"]],
        on="ticker",
        how="inner",
    )

    # Point-in-time float: for each (ticker, settlement_date), use the
    # most recent CapIQ float observation on or before that date.
    # merge_asof requires the `on` key globally sorted on both sides.
    if not floats.empty:
        df = df.sort_values("settlement_date").reset_index(drop=True)
        floats_sorted = (
            floats.rename(columns={"asof_date": "settlement_date"})
                  .sort_values("settlement_date")
                  .reset_index(drop=True)
        )
        df = pd.merge_asof(
            df,
            floats_sorted,
            on="settlement_date",
            by="ticker",
            direction="backward",
        )
    else:
        df["float_shares"] = pd.NA

    df["si_pct_float"] = df["si_shares"] / df["float_shares"]
    # cap at sane max so junk denominators don't dominate ranks
    df.loc[df["si_pct_float"] > 1.5, "si_pct_float"] = pd.NA

    return df.reset_index(drop=True)


if __name__ == "__main__":
    # Quick smoke test
    import time

    t0 = time.time()
    si = load_si_history()
    print(f"SI history:    {len(si):>10,} rows  ({time.time()-t0:.1f}s)")

    t0 = time.time()
    eq = load_equities()
    print(f"Equities (US): {len(eq):>10,} rows  ({time.time()-t0:.1f}s)")

    t0 = time.time()
    fl = load_float_panel()
    print(f"Float panel:   {len(fl):>10,} rows  ({time.time()-t0:.1f}s)  unique tickers: {fl['ticker'].nunique() if len(fl) else 0:,}")

    t0 = time.time()
    j = load_joined_si_with_meta(si, eq, fl)
    print(f"Joined:        {len(j):>10,} rows  ({time.time()-t0:.1f}s)  unique tickers: {j['ticker'].nunique():,}")
    print(f"  with float:  {j['float_shares'].notna().sum():>10,} rows  ({j[j['float_shares'].notna()]['ticker'].nunique():,} unique tickers)")
    print(f"  date range:  {j['settlement_date'].min()} → {j['settlement_date'].max()}")
    print(f"\nSample rows:")
    print(j[j["ticker"].isin(["BRBR", "AAPL", "GME", "NVDA"])]
          .sort_values(["ticker", "settlement_date"])
          .tail(8)
          [["ticker", "settlement_date", "si_shares", "float_shares", "si_pct_float", "sector", "market_class"]]
          .to_string(index=False))
