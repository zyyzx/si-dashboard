#!/usr/bin/env python3
"""Filter the tickers from capiq_mcap_check.xlsx by primary exchange
(US-tradeable only) AND market cap threshold, writing the kept list to
tickers_above_<N>m.txt.

create_capiq_template.py auto-detects the resulting file and uses it as
the ticker universe instead of every ticker in si_dashboard.html.

Workbook layout supported:
  - New (3 columns): Ticker, Market Cap USD ($M), Primary Exchange
  - Old (2 columns): Ticker, Market Cap ($M)  -- falls back to mcap-only
                     filtering with a sanity upper bound. Re-run
                     create_mcap_check.py and recalc in Excel to upgrade.

Run:  python filter_tickers_above_mcap.py [threshold_in_millions]
      (default: 1000  -> $1B)
"""

import os
import sys

import openpyxl


TRACKER_DIR = os.path.dirname(os.path.abspath(__file__))

# Substring tokens for US-tradeable exchange names. CapIQ returns
# variants like "NYSE", "NasdaqGS", "NasdaqGM", "NasdaqCM",
# "NYSEAM" / "NYSE American", "NYSEARCA" / "ARCA", "Cboe BZX" / "BATS".
# Matching by substring (case-insensitive) handles all of these without
# enumerating every form CapIQ might emit.
_US_EXCHANGE_TOKENS = [
    "NYSE",      # NYSE, NYSEAM, NYSEARCA, NYSE American
    "NASDAQ",    # NasdaqGS, NasdaqGM, NasdaqCM
    "AMEX",      # legacy AMEX label
    "ARCA",      # alternate NYSEARCA form
    "BATS",      # legacy BATS
    "BZX",       # Cboe BZX
    "CBOE",      # Cboe variants
    "IEX",       # IEX-listed names
]

# Defense in depth: even with USD-forced mcap, a residual class of
# tickers can return implausibly large values (CapIQ data quality bugs
# unrelated to currency). Cap at $5T -- NVDA is the current top of
# market at ~$4-4.5T, leaves ~25% headroom for real-mega-cap growth.
_SANE_MAX_MCAP_M = 5_000_000  # = $5 trillion


def _is_us_exchange(s) -> bool:
    """True if the exchange string looks like a US-tradeable venue."""
    if not s:
        return False
    s_up = str(s).upper()
    return any(tok in s_up for tok in _US_EXCHANGE_TOKENS)


def main() -> None:
    threshold_m = float(sys.argv[1]) if len(sys.argv) > 1 else 1000.0

    src = os.path.join(TRACKER_DIR, "capiq_mcap_check.xlsx")
    if not os.path.exists(src):
        raise SystemExit(
            f"missing {src} -- run create_mcap_check.py first, "
            "then open in Excel + CapIQ Pro, let recalc finish, save."
        )

    wb = openpyxl.load_workbook(src, data_only=True)
    if "MarketCap" not in wb.sheetnames:
        raise SystemExit("capiq_mcap_check.xlsx has no 'MarketCap' sheet -- "
                          "re-run create_mcap_check.py")
    ws = wb["MarketCap"]

    # Detect whether the workbook has the new 3-column layout (with
    # exchange) or the old 2-column one. Header row dictates.
    headers = [str(c.value or "").lower() for c in ws[1]]
    has_exchange_col = any("exchange" in h for h in headers)
    if has_exchange_col:
        print(f"Workbook layout: NEW (ticker / mcap-USD / exchange)")
    else:
        print(f"Workbook layout: OLD (ticker / mcap only -- exchange filter "
              f"unavailable; run create_mcap_check.py + recalc to upgrade)")

    kept: list[tuple[str, float, str]] = []
    no_value: list[str] = []
    below: list[tuple[str, float]] = []
    unit_bug: list[tuple[str, float]] = []      # implausibly large
    non_us_exch: list[tuple[str, str]] = []     # exchange not in whitelist
    for r in range(2, ws.max_row + 1):
        t = ws.cell(row=r, column=1).value
        mc = ws.cell(row=r, column=2).value
        exch = ws.cell(row=r, column=3).value if has_exchange_col else None
        if t is None:
            continue
        t = str(t).strip().upper()
        if has_exchange_col and not _is_us_exchange(exch):
            non_us_exch.append((t, str(exch) if exch else "(blank)"))
            continue
        if not isinstance(mc, (int, float)):
            no_value.append(t)
            continue
        mc_f = float(mc)
        if mc_f > _SANE_MAX_MCAP_M:
            unit_bug.append((t, mc_f))
            continue
        if mc_f >= threshold_m:
            kept.append((t, mc_f, str(exch) if exch else ""))
        else:
            below.append((t, mc_f))

    kept.sort(key=lambda x: -x[1])
    out_path = os.path.join(TRACKER_DIR, f"tickers_above_{int(threshold_m)}m.txt")
    with open(out_path, "w") as f:
        for t, _, _ in kept:
            f.write(t + "\n")

    total = (len(kept) + len(below) + len(no_value)
              + len(unit_bug) + len(non_us_exch))
    print(f"Threshold:        ${threshold_m:,.0f}M ({threshold_m/1000:.1f}B)")
    print(f"Sanity upper bound:${_SANE_MAX_MCAP_M:,.0f}M "
          f"(${_SANE_MAX_MCAP_M/1_000_000:.0f}T)")
    print(f"Kept:             {len(kept):,} / {total:,} tickers")
    print(f"Below threshold:  {len(below):,}")
    print(f"No mcap (NA):     {len(no_value):,}  "
          f"(CapIQ returned blank/error; usually delisted or pre-IPO)")
    if has_exchange_col:
        print(f"Non-US exchange:  {len(non_us_exch):,}  "
              f"(OTC pinks, foreign primaries -- not US-tradeable)")
    print(f"Unit bug (>${_SANE_MAX_MCAP_M/1_000_000:.0f}T):  {len(unit_bug):,}  "
          f"(residual data-quality outliers)")
    print(f"Output:           {out_path}")
    print()
    if kept:
        print(f"Sample of top 10 by mcap:")
        for t, mc, exch in kept[:10]:
            exch_label = f"  [{exch}]" if exch else ""
            print(f"  {t:6}  ${mc:>14,.0f}M{exch_label}")
    if non_us_exch and has_exchange_col:
        print(f"\nSample of non-US-exchange drops (first 10 of "
              f"{len(non_us_exch):,}):")
        for t, e in non_us_exch[:10]:
            print(f"  {t:6}  exchange={e}")
    if unit_bug:
        print(f"\nUnit-bug drops (first 5 of {len(unit_bug):,}):")
        unit_bug.sort(key=lambda x: -x[1])
        for t, mc in unit_bug[:5]:
            print(f"  {t:6}  raw_value={mc:,.0f}M")
    print(f"\nNext: python create_capiq_template.py")
    print(f"      (auto-detects {os.path.basename(out_path)} and uses it)")


if __name__ == "__main__":
    main()
