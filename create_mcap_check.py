#!/usr/bin/env python3
"""Generate a minimal CapIQ workbook to grab current market cap (in USD)
AND the primary listing exchange for every ticker in the dashboard. Used
as a pre-filter step before create_capiq_template.py.

Why two columns instead of one:
  * USD-forced mcap kills the local-currency-leak problem for foreign
    ADRs that came back as $4-8T in the millions column.
  * IQ_PRIMARY_EXCHANGE lets us drop non-US-tradeable tickers (OTC pinks,
    foreign primary listings) that you can't act on from a US brokerage
    anyway. The dashboard's marketClass field is empty in si_dashboard.html
    so we can't get exchange info from there.

After this recalcs, run filter_tickers_above_mcap.py to write the kept
list. Workbook size ~10-14k tickers * 2 formula columns = ~20-28k cells,
so recalc is still under 15 min on a typical CapIQ Pro tier.

Run:  python create_mcap_check.py
"""

import json
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill


TRACKER_DIR = os.path.dirname(os.path.abspath(__file__))


def main() -> None:
    print("Loading tickers from dashboard...")
    with open(os.path.join(TRACKER_DIR, "si_dashboard.html"), encoding="utf-8") as f:
        text = f.read()
    m = re.search(r"const RAW=(\{.*?\});", text)
    if not m:
        raise SystemExit("could not find RAW={...} in si_dashboard.html "
                          "-- regenerate the dashboard first")
    raw = json.loads(m.group(1))

    tickers = sorted(raw["tickers"].keys())
    print(f"Tickers: {len(tickers):,}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MarketCap"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="16213E")
    for col, label in enumerate(
        ("Ticker", "Market Cap USD ($M)", "Primary Exchange"), start=1
    ):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18

    # Two formulas per ticker.
    # B: USD-forced market cap. The 4th positional arg of CIQ() is the
    #    currency override; the 3rd is left blank to mean "current".
    # C: Primary exchange. filter_tickers_above_mcap.py whitelists
    #    NYSE / NASDAQ / NYSE American / ARCA / BZX strings (substring
    #    match, so variants like NasdaqGS / NasdaqCM all pass).
    for i, t in enumerate(tickers, start=2):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2,
                 value=f'=@CIQ($A{i},"IQ_MARKETCAP",,"USD")')
        ws.cell(row=i, column=3,
                 value=f'=@CIQ($A{i},"IQ_PRIMARY_EXCHANGE")')
    ws.freeze_panes = "B2"

    out = os.path.join(TRACKER_DIR, "capiq_mcap_check.xlsx")
    wb.save(out)
    print(f"\nSaved {out}")
    print(f"\nNext: open in Excel with CapIQ Pro, let recalc finish "
          f"(~10-15 min for {len(tickers)*2:,} formulas), save in place, then:")
    print(f"      python filter_tickers_above_mcap.py 1000")
    print(f"      (1000 = $1B floor; pass a different number to tune)")


if __name__ == "__main__":
    main()
