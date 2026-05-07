#!/usr/bin/env python3
"""
Integrate CapIQ float data into the SI dashboard.

Supports two formats:
  1. Historical template (capiq_float_historical.xlsx): tickers × quarterly dates
     - Interpolates float for bi-weekly settlement periods between quarterly snapshots
  2. Simple template (capiq_float_template.xlsx): single float value per ticker
  3. CSV (float_data.csv): ticker,float_shares

Usage:
    python integrate_float_data.py [--capiq capiq_float_historical.xlsx] [--dashboard si_dashboard.html]
"""

import csv
import json
import re
import sys
import datetime
from pathlib import Path

TRACKER_DIR = Path(__file__).parent
CAPIQ_HIST = TRACKER_DIR / "capiq_float_historical.xlsx"
CAPIQ_SIMPLE = TRACKER_DIR / "capiq_float_template.xlsx"
FLOAT_CSV = TRACKER_DIR / "float_data.csv"
DASHBOARD_FILE = TRACKER_DIR / "si_dashboard.html"


def load_historical_float_from_xlsx(path):
    """Load historical float data: returns {ticker: {date_str: float_shares}}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Float"]

    dates = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        if isinstance(val, datetime.datetime):
            dates.append(val.strftime("%Y%m%d"))
        elif isinstance(val, str):
            dates.append(val.replace("-", "").replace("/", ""))
        else:
            dates.append(str(val))

    float_data = {}
    skipped = 0

    for row in range(2, ws.max_row + 1):
        ticker = ws.cell(row=row, column=1).value
        if not ticker:
            continue
        ticker = str(ticker).strip()
        series = {}
        for j, d in enumerate(dates):
            val = ws.cell(row=row, column=j + 2).value
            if val is not None and isinstance(val, (int, float)) and val > 0:
                # CapIQ IQ_FLOAT returns values in millions; convert to raw shares
                series[d] = float(val) * 1e6
        if series:
            float_data[ticker] = series
        else:
            skipped += 1

    return float_data, skipped, dates


def load_simple_float_from_xlsx(path):
    """Load single-value float data: returns {ticker: float_shares}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Tickers"]
    float_data = {}
    skipped = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        ticker = row[0]
        if not ticker:
            continue
        ticker = str(ticker).strip()
        float_shares = row[1] if len(row) > 1 else None
        shares_out = row[2] if len(row) > 2 else None
        float_pct = row[3] if len(row) > 3 else None

        if float_shares and isinstance(float_shares, (int, float)) and float_shares > 0:
            float_data[ticker] = float(float_shares)
        elif shares_out and float_pct and isinstance(shares_out, (int, float)) and isinstance(float_pct, (int, float)):
            float_data[ticker] = float(shares_out) * float(float_pct) / 100.0
        elif shares_out and isinstance(shares_out, (int, float)) and shares_out > 0:
            float_data[ticker] = float(shares_out)
        else:
            skipped += 1

    return float_data, skipped


def interpolate_float(float_series, all_dates):
    """Given {date_str: float_val} at quarterly points, return {date_str: float_val} for all dates.

    Uses the most recent known float value (forward-fill / step interpolation).
    """
    sorted_known = sorted(float_series.keys())
    result = {}
    known_idx = 0

    for d in all_dates:
        while known_idx < len(sorted_known) - 1 and sorted_known[known_idx + 1] <= d:
            known_idx += 1
        if sorted_known[known_idx] <= d:
            result[d] = float_series[sorted_known[known_idx]]
        elif known_idx == 0:
            result[d] = float_series[sorted_known[0]]

    return result


def update_dashboard_historical(dashboard_path, float_data_hist):
    """Update dashboard using historical (time-varying) float data."""
    print(f"\nReading dashboard: {dashboard_path}")
    with open(dashboard_path, "r", encoding="utf-8") as f:
        html = f.read()

    m = re.search(r"const RAW=(\{.*?\});", html)
    if not m:
        print("ERROR: Could not find RAW data in dashboard HTML")
        sys.exit(1)

    raw = json.loads(m.group(1))
    tickers = raw["tickers"]
    all_dates = raw["dates"]

    updated = 0
    already_had = 0
    no_float = 0

    for sym, t in tickers.items():
        if sym not in float_data_hist:
            no_float += 1
            continue

        float_series = float_data_hist[sym]
        if not float_series:
            no_float += 1
            continue

        interpolated = interpolate_float(float_series, all_dates)
        had_pct = bool(t.get("pct") and len(t["pct"]) > 0)

        pct_series = []
        for idx, si_val in t["si"]:
            date_str = all_dates[idx]
            if date_str in interpolated and interpolated[date_str] > 0:
                pct_val = round((si_val / interpolated[date_str]) * 100, 4)
                pct_series.append([idx, pct_val])

        if pct_series:
            t["pct"] = pct_series
            if had_pct:
                already_had += 1
            updated += 1
        else:
            no_float += 1

    print(f"\nFloat integration results:")
    print(f"  Tickers with float data applied: {updated:,}")
    print(f"  - Previously had pct (overwritten): {already_had:,}")
    print(f"  - Newly added pct: {updated - already_had:,}")
    print(f"  Tickers without float data: {no_float:,}")

    new_json = json.dumps(raw, separators=(",", ":"))
    new_html = html[: m.start(1)] + new_json + html[m.end(1) :]

    size_before = len(html)
    size_after = len(new_html)
    print(f"\n  Dashboard size: {size_before/1e6:.1f}MB -> {size_after/1e6:.1f}MB")

    with open(dashboard_path, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"  Dashboard updated: {dashboard_path}")
    return updated


def update_dashboard_simple(dashboard_path, float_data):
    """Update dashboard using single float value per ticker."""
    print(f"\nReading dashboard: {dashboard_path}")
    with open(dashboard_path, "r", encoding="utf-8") as f:
        html = f.read()

    m = re.search(r"const RAW=(\{.*?\});", html)
    if not m:
        print("ERROR: Could not find RAW data in dashboard HTML")
        sys.exit(1)

    raw = json.loads(m.group(1))
    tickers = raw["tickers"]

    updated = 0
    already_had = 0
    no_float = 0

    for sym, t in tickers.items():
        if sym not in float_data:
            no_float += 1
            continue

        float_shares = float_data[sym]
        if float_shares <= 0:
            no_float += 1
            continue

        had_pct = bool(t.get("pct") and len(t["pct"]) > 0)

        pct_series = []
        for idx, si_val in t["si"]:
            pct_val = round((si_val / float_shares) * 100, 4)
            pct_series.append([idx, pct_val])

        t["pct"] = pct_series
        if had_pct:
            already_had += 1
        updated += 1

    print(f"\nFloat integration results:")
    print(f"  Tickers with float data applied: {updated:,}")
    print(f"  - Previously had pct (overwritten): {already_had:,}")
    print(f"  - Newly added pct: {updated - already_had:,}")
    print(f"  Tickers without float data: {no_float:,}")

    new_json = json.dumps(raw, separators=(",", ":"))
    new_html = html[: m.start(1)] + new_json + html[m.end(1) :]

    size_before = len(html)
    size_after = len(new_html)
    print(f"\n  Dashboard size: {size_before/1e6:.1f}MB -> {size_after/1e6:.1f}MB")

    with open(dashboard_path, "w", encoding="utf-8") as f:
        f.write(new_html)

    print(f"  Dashboard updated: {dashboard_path}")
    return updated


def main():
    import argparse

    parser = argparse.ArgumentParser(description="Integrate float data into SI dashboard")
    parser.add_argument("--capiq", default=None, help="Path to filled CapIQ Excel template")
    parser.add_argument("--csv", default=None, help="Path to float CSV (ticker,float_shares)")
    parser.add_argument("--dashboard", default=str(DASHBOARD_FILE), help="Path to dashboard HTML")
    args = parser.parse_args()

    print("=" * 60)
    print("INTEGRATING FLOAT DATA INTO SI DASHBOARD")
    print("=" * 60)

    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Run: pip install openpyxl")
        sys.exit(1)

    dashboard_path = Path(args.dashboard)
    if not dashboard_path.exists():
        print(f"ERROR: Dashboard not found: {dashboard_path}")
        sys.exit(1)

    capiq_path = Path(args.capiq) if args.capiq else None

    # Try historical template first, then simple, then CSV
    if capiq_path and capiq_path.exists():
        wb = openpyxl.load_workbook(capiq_path, data_only=True)
        if "Float" in wb.sheetnames:
            print(f"\nLoading HISTORICAL float data from: {capiq_path}")
            float_data, skipped, template_dates = load_historical_float_from_xlsx(capiq_path)
            print(f"  Loaded {len(float_data):,} tickers with time-series float (skipped {skipped:,})")
            print(f"  Template dates: {len(template_dates)}")
            updated = update_dashboard_historical(dashboard_path, float_data)
        elif "Tickers" in wb.sheetnames:
            print(f"\nLoading SIMPLE float data from: {capiq_path}")
            float_data, skipped = load_simple_float_from_xlsx(capiq_path)
            print(f"  Loaded {len(float_data):,} tickers (skipped {skipped:,})")
            updated = update_dashboard_simple(dashboard_path, float_data)
        else:
            print(f"ERROR: Unrecognized template format in {capiq_path}")
            sys.exit(1)
        wb.close()
    elif CAPIQ_HIST.exists():
        print(f"\nLoading HISTORICAL float data from: {CAPIQ_HIST}")
        float_data, skipped, template_dates = load_historical_float_from_xlsx(CAPIQ_HIST)
        print(f"  Loaded {len(float_data):,} tickers (skipped {skipped:,})")
        updated = update_dashboard_historical(dashboard_path, float_data)
    elif CAPIQ_SIMPLE.exists():
        print(f"\nLoading SIMPLE float data from: {CAPIQ_SIMPLE}")
        float_data, skipped = load_simple_float_from_xlsx(CAPIQ_SIMPLE)
        print(f"  Loaded {len(float_data):,} tickers (skipped {skipped:,})")
        updated = update_dashboard_simple(dashboard_path, float_data)
    elif args.csv:
        csv_path = Path(args.csv)
        print(f"\nLoading float data from CSV: {csv_path}")
        float_data = {}
        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                ticker = row.get("ticker", "").strip()
                val = row.get("float_shares", "")
                if ticker and val:
                    try:
                        float_data[ticker] = float(val)
                    except ValueError:
                        pass
        print(f"  Loaded {len(float_data):,} tickers")
        updated = update_dashboard_simple(dashboard_path, float_data)
    else:
        print(f"\nERROR: No float data source found.")
        print(f"  Expected one of:")
        print(f"    {CAPIQ_HIST}")
        print(f"    {CAPIQ_SIMPLE}")
        print(f"  Or pass --capiq <path> or --csv <path>")
        sys.exit(1)

    print(f"\n{'=' * 60}")
    print(f"DONE. {updated:,} tickers now have SI % of Float data.")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
