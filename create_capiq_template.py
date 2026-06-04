#!/usr/bin/env python3
"""Create CapIQ historical float template with CIQ formulas.

By default the workbook covers every ticker in si_dashboard.html across
the dates in quarterly_dates.json. When a tickers_above_<N>m.txt file is
present (produced by filter_tickers_above_mcap.py), the script narrows
the universe to that filtered list instead -- typically cutting the
formula count by ~80% by dropping sub-$1B names.

Pre-filter workflow:
    python create_mcap_check.py            # tiny mcap workbook
    # open + Excel/CapIQ recalc + save
    python filter_tickers_above_mcap.py 1000   # writes tickers_above_1000m.txt
    python create_capiq_template.py        # slim float workbook

Run as-is for the full universe (no pre-filter):
    python create_capiq_template.py
"""

import datetime
import glob
import json
import os
import re

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


TRACKER_DIR = os.path.dirname(os.path.abspath(__file__))


def _load_filter_list() -> tuple[list[str] | None, str | None]:
    """Return (tickers, source_path) when a tickers_above_<N>m.txt exists in
    TRACKER_DIR; otherwise (None, None). If multiple thresholds are present
    we pick the highest (most-restrictive) since the user clearly ran the
    filter most recently with that threshold."""
    candidates = sorted(
        glob.glob(os.path.join(TRACKER_DIR, "tickers_above_*m.txt")),
        key=lambda p: int(re.search(r"tickers_above_(\d+)m\.txt$", p).group(1))
                       if re.search(r"tickers_above_(\d+)m\.txt$", p) else 0,
        reverse=True,
    )
    if not candidates:
        return None, None
    chosen = candidates[0]
    with open(chosen) as f:
        tickers = [line.strip().upper() for line in f if line.strip()]
    return tickers, chosen


print("Loading tickers from dashboard...")
with open(os.path.join(TRACKER_DIR, "si_dashboard.html"), "r", encoding="utf-8") as f:
    text = f.read()
m = re.search(r"const RAW=(\{.*?\});", text)
raw = json.loads(m.group(1))
all_tickers_from_dash = sorted(raw["tickers"].keys())
all_dates = raw["dates"]

filtered, filter_src = _load_filter_list()
if filtered:
    # Intersect so we only request tickers the dashboard actually knows about
    # (a stale filter list won't introduce ghosts).
    dash_set = set(all_tickers_from_dash)
    tickers = sorted([t for t in filtered if t in dash_set])
    dropped = len(filtered) - len(tickers)
    print(f"Filter applied: {os.path.basename(filter_src)} "
          f"-> {len(tickers):,} tickers "
          f"({dropped:,} from the filter weren't in si_dashboard.html and were dropped)")
else:
    tickers = all_tickers_from_dash
    print(f"No tickers_above_*m.txt filter found -- using all "
          f"{len(tickers):,} tickers from si_dashboard.html")
    print(f"  (run create_mcap_check.py + filter_tickers_above_mcap.py first "
          f"to narrow this list)")

with open(os.path.join(TRACKER_DIR, "quarterly_dates.json")) as f:
    selected_dates = json.load(f)

print(f"Tickers: {len(tickers):,}, Dates: {len(selected_dates)}")

wb = openpyxl.Workbook()

# Instructions sheet
ws_inst = wb.active
ws_inst.title = "Instructions"
lines = [
    "CapIQ Historical Float Data Template",
    "",
    "Formulas: =@CIQ($A{row}, \"IQ_FLOAT\", {col}$1)",
    "Each cell pulls the public float for a ticker as of the settlement date in row 1.",
    "",
    "Steps:",
    "1. Open this file in Excel with the Capital IQ Pro Plug-in",
    "2. Go to the Float sheet - formulas are pre-filled",
    f"3. Let CIQ calculate (~{len(tickers)*len(selected_dates):,} formulas, may take 30-60 min)",
    "4. Save As a new copy (values only recommended)",
    "5. Run: python integrate_float_data.py --capiq capiq_float_historical.xlsx",
    "",
    f"Layout: {len(tickers):,} tickers x {len(selected_dates)} quarterly dates",
    f"Dates: {selected_dates[0]} through {selected_dates[-1]}",
    f"Total formulas: {len(tickers) * len(selected_dates):,}",
]
if filter_src:
    lines.insert(2, f"Filter source: {os.path.basename(filter_src)} "
                     f"(market-cap pre-filter)")
for i, line in enumerate(lines, 1):
    ws_inst.cell(row=i, column=1, value=line)
ws_inst["A1"].font = Font(bold=True, size=14, color="0000FF")
ws_inst.column_dimensions["A"].width = 80

# Float sheet
print("Building Float sheet with CIQ formulas...")
ws = wb.create_sheet("Float")

header_fill = PatternFill("solid", fgColor="16213E")
header_font = Font(bold=True, color="FFFFFF", size=9)

ws.cell(row=1, column=1, value="Ticker").font = header_font
ws["A1"].fill = header_fill
ws.column_dimensions["A"].width = 10

for j, dstr in enumerate(selected_dates):
    col = j + 2
    dt = datetime.datetime.strptime(dstr, "%Y%m%d")
    cell = ws.cell(row=1, column=col, value=dt)
    cell.number_format = "MM/DD/YYYY"
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(col)].width = 14

for i, ticker in enumerate(tickers):
    row = i + 2
    ws.cell(row=row, column=1, value=ticker)
    for j in range(len(selected_dates)):
        col = j + 2
        col_letter = get_column_letter(col)
        formula = f'=@CIQ($A{row},"IQ_FLOAT",{col_letter}$1)'
        ws.cell(row=row, column=col, value=formula)
    if (i + 1) % 2000 == 0:
        print(f"  {i+1:,}/{len(tickers):,} tickers...")

ws.freeze_panes = "B2"

# Reference sheet with all settlement dates
ws_ref = wb.create_sheet("All Settlement Dates")
ws_ref.cell(row=1, column=1, value="Date Index").font = Font(bold=True)
ws_ref.cell(row=1, column=2, value="Settlement Date").font = Font(bold=True)
ws_ref.cell(row=1, column=3, value="In Template").font = Font(bold=True)
for i, d in enumerate(all_dates):
    dt = datetime.datetime.strptime(d, "%Y%m%d")
    ws_ref.cell(row=i + 2, column=1, value=i)
    ws_ref.cell(row=i + 2, column=2, value=dt).number_format = "MM/DD/YYYY"
    ws_ref.cell(row=i + 2, column=3, value="Yes" if d in selected_dates else "")
ws_ref.column_dimensions["A"].width = 12
ws_ref.column_dimensions["B"].width = 16
ws_ref.column_dimensions["C"].width = 14

out_path = os.path.join(TRACKER_DIR, "capiq_float_historical.xlsx")
print(f"Saving to {out_path}...")
wb.save(out_path)
print(f"Done! {len(tickers):,} tickers x {len(selected_dates)} dates = "
      f"{len(tickers)*len(selected_dates):,} formulas")
